"""
Shared logic for WineMore Reports — inventory reporting and import CSV generation.
Works with in-memory file objects (Streamlit uploads) instead of file paths.
"""

import csv
import io
import math
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── helpers ────────────────────────────────────────────────────────────────

def to_int(val):
    v = (val or "").strip()
    if not v or v.lower() == "not stocked":
        return 0
    try:
        return int(float(v))
    except Exception:
        return 0

def to_float(val):
    try:
        return float((val or "").strip())
    except Exception:
        return 0.0


# ─── Joval rules ─────────────────────────────────────────────────────────────
#
# 1. Quantity = floor(Qty_in_stock × carton_size)
# 2. Price < $50   → zero if qty < 30
#    Price $50–200 → zero if qty < 10
#    Price > $200  → zero if qty < 5
# 3. Any supplier-joval SKU absent from the SOH file → zero

def _parse_carton_size(raw):
    """Parse "6/ctn" or "12/ctn" → int. Returns 1 if unparseable."""
    try:
        return int(str(raw).strip().split("/")[0])
    except Exception:
        return 1

def apply_joval_threshold(qty, price):
    """Return final qty after applying price-tier zero thresholds."""
    if price < 50 and qty < 30:
        return 0
    if 50 <= price <= 200 and qty < 10:
        return 0
    if price > 200 and qty < 5:
        return 0
    return qty


# ─── loaders ────────────────────────────────────────────────────────────────

def _read_file(f):
    """Read an uploaded file, seeking to 0 first. Returns decoded text."""
    if hasattr(f, "seek"):
        f.seek(0)
    raw = f.read()
    return raw.decode("utf-8") if isinstance(raw, bytes) else raw

def _is_xlsx(f):
    name = getattr(f, "name", "") or ""
    return name.lower().endswith(".xlsx")

def _iter_rows(f):
    """
    Yield row dicts from a CSV or XLSX file.
    For XLSX: row 1 is the header, remaining rows become dicts.
    Handles Streamlit UploadedFile objects and regular file-like objects.
    """
    if _is_xlsx(f):
        if hasattr(f, "seek"):
            f.seek(0)
        buf = io.BytesIO(f.read())
        wb  = openpyxl.load_workbook(buf, data_only=True)
        ws  = wb.active
        headers = [str(ws.cell(1, c).value or "").strip()
                   for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            row = {headers[c]: str(ws.cell(r, c + 1).value or "").strip()
                   for c in range(len(headers))}
            # skip entirely blank rows
            if any(v for v in row.values()):
                yield row
    else:
        yield from csv.DictReader(io.StringIO(_read_file(f)))

def load_products_from_files(files):
    """Return {handle: {title, price, tags, skus}} from products_export CSV/XLSX.
    Accepts 1 or 2 files — Shopify splits large exports across multiple files."""
    products = {}
    for f in files:
        current_handle = current_title = ""
        for row in _iter_rows(f):
            h = row.get("Handle", "").strip()
            if row.get("Title", "").strip():
                current_handle = h
                current_title = row["Title"].strip()
            if not current_handle:
                continue
            if current_handle not in products:
                products[current_handle] = {
                    "title": current_title,
                    "price": row.get("Variant Price", "").strip(),
                    "tags":  row.get("Tags", "").strip(),
                    "skus":  [],
                }
            sku = row.get("Variant SKU", "").strip()
            if sku and sku not in products[current_handle]["skus"]:
                products[current_handle]["skus"].append(sku)
    return products

def load_sku_prices(files):
    """Return {sku: float_price} from products_export CSV/XLSX (1 or 2 files)."""
    prices = {}
    for f in files:
        for row in _iter_rows(f):
            sku = row.get("Variant SKU", "").strip()
            if sku and sku not in prices:
                prices[sku] = to_float(row.get("Variant Price", ""))
    return prices

def load_sku_titles(files):
    """Return {sku: product_title} from products_export CSV/XLSX (1 or 2 files)."""
    titles = {}
    current_title = ""
    for f in files:
        for row in _iter_rows(f):
            if row.get("Title", "").strip():
                current_title = row["Title"].strip()
            sku = row.get("Variant SKU", "").strip()
            if sku and sku not in titles:
                titles[sku] = current_title
    return titles

def load_supplier_skus(files, tag="supplier-joval"):
    """Return set of SKUs whose product has the given supplier tag (1 or 2 files)."""
    tagged = set()
    for f in files:
        current_skus = []
        current_tagged = False
        for row in _iter_rows(f):
            if row.get("Title", "").strip():
                current_skus = []
                current_tagged = tag in row.get("Tags", "")
            sku = row.get("Variant SKU", "").strip()
            if sku:
                current_skus.append(sku)
            if current_tagged:
                tagged.update(current_skus)
    return tagged

def load_inventory_from_files(files):
    """Return {handle: {wmc, chad, skus, title, _rows}} from inventory_export CSV/XLSX."""
    inv = {}
    for f in files:
        for row in _iter_rows(f):
            h   = row.get("Handle", "").strip()
            sku = row.get("SKU", "").strip()
            if not h:
                continue
            wmc   = to_int(row.get("Wine More Cellars", ""))
            chad  = to_int(row.get("WM Chadstone", ""))
            title = row.get("Title", "").strip()
            if h not in inv:
                inv[h] = {"wmc": 0, "chad": 0, "skus": [], "title": title, "_rows": []}
            inv[h]["wmc"]  += wmc
            inv[h]["chad"] += chad
            if sku:
                inv[h]["skus"].append(sku)
            inv[h]["_rows"].append(row)
    return inv

def load_supplier_handles_from_csv(file):
    """Return set of handles/SKUs from a supplier CSV or XLSX."""
    handles = set()
    for row in _iter_rows(file):
        h = row.get("Handle", row.get("Item No.", "")).strip()
        if h:
            handles.add(h)
    return handles

def load_joval_quantities_xlsx(file, sku_prices=None, sku_titles=None):
    """
    Read Joval SOH xlsx — headers row 8, data from row 9:
      col A = Item No. (SKU)
      col D = Carton Size ("6/ctn")
      col G = Qty. In Stock (decimal cartons)

    Returns (qty_dict, joval_stats) where joval_stats contains breakdown of
    zeroed SKUs by reason (rules vs raw zero stock).
    """
    if hasattr(file, "seek"):
        file.seek(0)
    buf = io.BytesIO(file.read())
    wb = openpyxl.load_workbook(buf, data_only=True)
    ws = wb.active
    qty = {}
    zeroed_by_rules = []  # {"sku", "calculated", "price", "tier"}
    zeroed_raw      = []  # SKUs with 0 stock in SOH

    for r in range(9, ws.max_row + 1):
        sku    = ws.cell(r, 1).value
        carton = ws.cell(r, 4).value
        q      = ws.cell(r, 7).value
        if not sku:
            continue
        sku = str(sku).strip()
        carton_size = _parse_carton_size(carton)
        try:
            raw_qty = float(q) if q is not None else 0.0
        except (ValueError, TypeError):
            raw_qty = 0.0
        calculated = math.floor(raw_qty * carton_size)

        if calculated == 0:
            zeroed_raw.append(sku)

        if sku_prices and calculated > 0:
            price = sku_prices.get(sku, 0.0)
            final = apply_joval_threshold(calculated, price)
            if final == 0:
                tier = "Under $50" if price < 50 else ("$50–$200" if price <= 200 else "Over $200")
                zeroed_by_rules.append({
                    "SKU": sku,
                    "Product": sku_titles.get(sku, "") if sku_titles else "",
                    "Calculated Qty": calculated,
                    "Price": f"${price:.2f}",
                    "Tier": tier,
                })
            calculated = final

        qty[sku] = max(0, calculated)

    joval_stats = {
        "total":           len(qty),
        "nonzero":         sum(1 for v in qty.values() if v > 0),
        "zeroed_by_rules": zeroed_by_rules,
        "zeroed_raw":      zeroed_raw,
    }
    return qty, joval_stats

def load_qty_csv(file):
    """Generic CSV or XLSX with SKU and Quantity columns."""
    rows = list(_iter_rows(file))
    if not rows:
        return {}, "File is empty"
    fields_lower = [c.strip().lower() for c in rows[0].keys()]
    orig = list(rows[0].keys())
    sku_col = next((orig[i] for i, c in enumerate(fields_lower)
                    if c in ("sku", "item no.", "item no", "item_no")), None)
    qty_col = next((orig[i] for i, c in enumerate(fields_lower)
                    if c in ("quantity", "qty", "qty. in stock", "stock", "on hand")), None)
    if not sku_col or not qty_col:
        return {}, f"Cannot find SKU/Quantity columns (found: {orig})"
    qty = {}
    for row in rows:
        sku = row.get(sku_col, "").strip()
        q   = row.get(qty_col, "").strip()
        if sku and q:
            try:
                qty[sku] = max(0, int(float(q)))
            except (ValueError, TypeError):
                pass
    return qty, None

def load_flat_inventory_rows(files):
    """Return flat list of all raw row dicts from inventory exports (CSV or XLSX)."""
    rows = []
    for f in files:
        for row in _iter_rows(f):
            if row.get("SKU", "").strip():
                rows.append(row)
    return rows

def load_flat_rows_from_products(files):
    """
    Return flat list of variant row dicts from products export (1 or 2 files).
    Maps 'Variant SKU' → 'SKU' so build_inventory_import can consume them.
    Handles continuation rows (blank Title) correctly.
    """
    rows = []
    for f in files:
        current_handle = current_title = ""
        current_opts = {}
        for row in _iter_rows(f):
            h = row.get("Handle", "").strip()
            if row.get("Title", "").strip():
                current_handle = h
                current_title  = row["Title"].strip()
                current_opts   = {
                    "Option1 Name":  row.get("Option1 Name", ""),
                    "Option1 Value": row.get("Option1 Value", ""),
                    "Option2 Name":  row.get("Option2 Name", ""),
                    "Option2 Value": row.get("Option2 Value", ""),
                    "Option3 Name":  row.get("Option3 Name", ""),
                    "Option3 Value": row.get("Option3 Value", ""),
                }
            sku = row.get("Variant SKU", "").strip()
            if not sku or not current_handle:
                continue
            out = {"Handle": current_handle, "Title": current_title, "SKU": sku}
            out.update(current_opts)
            out["Option1 Value"] = row.get("Option1 Value", current_opts.get("Option1 Value", ""))
            out["Option2 Value"] = row.get("Option2 Value", current_opts.get("Option2 Value", ""))
            out["Option3 Value"] = row.get("Option3 Value", current_opts.get("Option3 Value", ""))
            rows.append(out)
    return rows


# ─── report builder ──────────────────────────────────────────────────────────

SUPPLIER_COLORS = {
    "joval":    "FFF2CC",
    "bibendum": "E2EFDA",
    "dws":      "DDEBF7",
    "sss":      "FCE4D6",
}

def build_report(products, inventory, supplier_map):
    """Return (openpyxl.Workbook, stats_dict) with Inventory + Summary sheets."""
    all_handles = set(products) | set(inventory)
    rows = []
    for h in all_handles:
        inv  = inventory.get(h, {"wmc": 0, "chad": 0, "skus": [], "title": ""})
        prod = products.get(h, {})
        wmc  = inv["wmc"]
        chad = inv["chad"]
        total = wmc + chad
        skus  = inv.get("skus", [])
        title = prod.get("title") or inv.get("title", "")

        if wmc > 0 and chad > 0:
            location = "Both"
        elif wmc > 0:
            location = "Wine More Cellars"
        elif chad > 0:
            location = "WM Chadstone"
        else:
            location = "—"

        rows.append({
            "SKU":               skus[0] if skus else "",
            "Name":              title,
            "Price":             to_float(prod.get("price", "")),
            "Wine More Cellars": wmc,
            "WM Chadstone":      chad,
            "Total Stock":       total,
            "Location":          location,
            "Supplier":          supplier_map.get(h, "—"),
            "zero":              total == 0,
        })
    rows.sort(key=lambda r: (r["zero"], r["Name"].lower()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    col_keys   = ["SKU", "Name", "Price", "Wine More Cellars", "WM Chadstone",
                  "Total Stock", "Location", "Supplier"]
    col_widths = [18, 55, 10, 20, 16, 14, 22, 20]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    zero_fill   = PatternFill("solid", fgColor="F2F2F2")
    grey_font   = Font(color="909090", size=10)
    normal_font = Font(size=10)
    thin_border = Border(bottom=Side(style="thin", color="D0D0D0"))

    ws.append(col_keys)
    for i in range(1, len(col_keys) + 1):
        c = ws.cell(1, i)
        c.fill      = header_fill
        c.font      = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(col_keys))}1"

    for r in rows:
        ws.append([
            r["SKU"], r["Name"],
            r["Price"] if r["Price"] else "",
            r["Wine More Cellars"], r["WM Chadstone"], r["Total Stock"],
            r["Location"], r["Supplier"],
        ])
        rn  = ws.max_row
        sup = r["Supplier"].split(",")[0].strip().lower()
        sup_color = SUPPLIER_COLORS.get(sup)
        for ci in range(1, len(col_keys) + 1):
            cell = ws.cell(rn, ci)
            cell.border = thin_border
            if r["zero"]:
                cell.fill = zero_fill
                cell.font = grey_font
            else:
                cell.font = normal_font
                if ci == len(col_keys) and sup_color:
                    cell.fill = PatternFill("solid", fgColor=sup_color)
        ws.cell(rn, 3).number_format = '"$"#,##0.00'

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Summary")
    stocked    = [r for r in rows if not r["zero"]]
    zeroed     = [r for r in rows if r["zero"]]
    total_wmc  = sum(r["Wine More Cellars"] for r in rows)
    total_chad = sum(r["WM Chadstone"] for r in rows)

    sup_stats = defaultdict(lambda: {"variants": 0, "stocked": 0, "zero": 0, "units": 0})
    for r in rows:
        for s in r["Supplier"].split(","):
            s = s.strip() or "No Supplier"
            sup_stats[s]["variants"] += 1
            sup_stats[s]["units"]    += r["Total Stock"]
            sup_stats[s]["stocked" if not r["zero"] else "zero"] += 1

    summary = [
        ["OVERVIEW", ""],
        ["Report generated", datetime.now().strftime("%d %b %Y %H:%M")],
        ["Total variants",   len(rows)],
        ["Stocked",          len(stocked)],
        ["Zero stock",       len(zeroed)],
        ["Total WMC units",  total_wmc],
        ["Total Chadstone",  total_chad],
        ["Combined units",   total_wmc + total_chad],
        [],
        ["SUPPLIER BREAKDOWN", "Variants", "Stocked", "Zero Stock", "Units"],
    ]
    for s, d in sorted(sup_stats.items(), key=lambda x: -x[1]["units"]):
        summary.append([s, d["variants"], d["stocked"], d["zero"], d["units"]])

    for row in summary:
        ws2.append(row)
    ws2["A1"].font = Font(bold=True, size=11)
    ws2["A10"].font = Font(bold=True)
    for col in ["A", "B", "C", "D", "E"]:
        ws2.column_dimensions[col].width = 18

    stats = {
        "total":      len(rows),
        "stocked":    len(stocked),
        "zero":       len(zeroed),
        "wmc_units":  total_wmc,
        "chad_units": total_chad,
    }
    return wb, stats


# ─── inventory import CSV builder ────────────────────────────────────────────

SHOPIFY_COLS = [
    "Handle", "Title", "Option1 Name", "Option1 Value",
    "Option2 Name", "Option2 Value", "Option3 Name", "Option3 Value",
    "SKU", "HS Code", "COO", "Location", "Bin name",
    "Incoming (not editable)", "Unavailable (not editable)",
    "Committed (not editable)", "Available (not editable)", "On hand (new)",
]

def build_inventory_import(inv_rows_raw, supplier_qty,
                           joval_skus=None, location="Wine More Cellars"):
    """
    Build Shopify inventory import CSV.

    inv_rows_raw  : flat list of raw dicts from product/inventory export
    supplier_qty  : {sku: int_qty} — already has Joval rules applied
    joval_skus    : set of all supplier-joval SKUs in Shopify — any absent
                    from supplier_qty get forced to 0 (rule 3)
    """
    zeroed_missing_skus = []
    if joval_skus:
        missing = joval_skus - set(supplier_qty.keys())
        zeroed_missing_skus = sorted(missing)
        for sku in missing:
            supplier_qty[sku] = 0

    matched   = []
    skipped   = 0
    not_found = set(supplier_qty.keys())

    for row in inv_rows_raw:
        sku = row.get("SKU", "").strip()
        if sku not in supplier_qty:
            skipped += 1
            continue
        not_found.discard(sku)
        qty_val = supplier_qty[sku]
        out = {col: "" for col in SHOPIFY_COLS}
        out.update({
            "Handle":                     row.get("Handle", ""),
            "Title":                      row.get("Title", ""),
            "Option1 Name":               row.get("Option1 Name", ""),
            "Option1 Value":              row.get("Option1 Value", ""),
            "Option2 Name":               row.get("Option2 Name", ""),
            "Option2 Value":              row.get("Option2 Value", ""),
            "Option3 Name":               row.get("Option3 Name", ""),
            "Option3 Value":              row.get("Option3 Value", ""),
            "SKU":                        sku,
            "Location":                   location,
            "Incoming (not editable)":    "0",
            "Unavailable (not editable)": "0",
            "Committed (not editable)":   "0",
            "Available (not editable)":   "0",
            "On hand (new)":              str(qty_val),
        })
        matched.append(out)

    zeroed  = sum(1 for r in matched if r["On hand (new)"] == "0")
    nonzero = len(matched) - zeroed

    preview = [
        {"SKU": r["SKU"], "Product": r.get("Title", ""),
         "Option": r["Option1 Value"], "On hand (new)": int(r["On hand (new)"])}
        for r in matched[:30]
    ]

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=SHOPIFY_COLS)
    writer.writeheader()
    writer.writerows(matched)

    stats = {
        "matched":             len(matched),
        "nonzero":             nonzero,
        "zeroed":              zeroed,
        "skipped":             skipped,
        "not_found":           len(not_found),
        "zeroed_missing_skus": zeroed_missing_skus,
        "preview":             preview,
    }
    return buf.getvalue().encode("utf-8"), stats, sorted(not_found)
